import openpyxl

wb = openpyxl.load_workbook("0805.xlsx")

ws_hello = wb["hello"]
ws_hello2 = wb["hello2"]

for i in range(1, 9):
	for n in range(1,3):
		copydata = ws_hello.cell(row = i, column = n)
		ws_hello2.cell(row = i, column = n).value = copydata.value

wb.save("0805.xlsx") 
