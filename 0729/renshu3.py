import openpyxl
wb = openpyxl.load_workbook("0729.xlsx")
ws = wb["Sheet"]
gokei = 0
for i in range(0,10):
	n = i + 2
	score = ws.cell(row = n, column = 2).value
	gokei = gokei + score
print(gokei)
average = gokei / 10
print(average)

ws.cell(row = 13, column = 1).value = "合計"
ws.cell(row = 13, column = 2).value = gokei
ws.cell(row = 14, column = 1).value = "平均"
ws.cell(row = 14, column = 2).value = average

wb.save("0729.xlsx")